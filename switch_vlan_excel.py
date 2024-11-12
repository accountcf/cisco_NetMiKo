from openpyxl import load_workbook
from netmiko import ConnectHandler
from getpass import getpass

# Nhập thông tin từ người dùng
switch_ip = input("Nhập IP của switch: ")
username = input("Nhập username: ")
password = getpass("Nhập password: ")

# Đọc file Excel
wb = load_workbook(filename='duong_dan_den_file_excel.xlsx')
sheet = wb.active

# Tạo kết nối đến switch
connection = ConnectHandler(device_type='cisco_ios', ip=switch_ip, username=username, password=password)

# Duyệt qua từng dòng trong file Excel và gửi lệnh
for row in sheet.iter_rows(min_row=2, values_only=True):  # Bắt đầu từ dòng thứ 2 để bỏ qua tiêu đề
    port, vlan = row
    # Tạo lệnh cấu hình
    config_commands = [
        f'interface {port}',
        f'switchport access vlan {vlan}',
        'switchport mode access',  # Bạn có thể thêm các lệnh khác nếu cần
        'no shutdown'
    ]
    # Gửi lệnh đến switch
    output = connection.send_config_set(config_commands)
    print(output)  # In kết quả ra màn hình để kiểm tra

# Đóng kết nối
connection.disconnect()

print("Đã cấu hình xong các cổng theo file Excel.")