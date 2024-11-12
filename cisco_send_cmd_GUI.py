import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from netmiko import ConnectHandler
from getpass import getpass
from tkinter import messagebox

def connect_and_configure():
    switch_ip = ip_entry.get()
    username = username_entry.get()
    password = password_entry.get()
    
    # Đọc file Excel đã chọn
    wb = load_workbook(filename=file_path.get())
    sheet = wb.active
    
    # Tạo kết nối đến switch
    connection = ConnectHandler(device_type='cisco_ios', ip=switch_ip, username=username, password=password)
    
    # Duyệt qua từng dòng trong file Excel và gửi lệnh
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Bỏ qua tiêu đề
        port, vlan = row
        # Tạo lệnh cấu hình
        config_commands = [
            f'interface {port}',
            f'switchport access vlan {vlan}',
            'switchport mode access',
            'no shutdown'
        ]
        # Gửi lệnh đến switch
        output = connection.send_config_set(config_commands)
        print(output)  # In kết quả ra màn hình để kiểm tra
    
    # Đóng kết nối
    connection.disconnect()
    
    messagebox.showinfo("Hoàn tất", "Đã cấu hình xong các cổng theo file Excel.")

def browse_file():
    filename = filedialog.askopenfilename()
    file_path.set(filename)

# Tạo cửa sổ chính
root = tk.Tk()
root.title("Cấu hình Switch Cisco")

# Tạo biến lưu đường dẫn file
file_path = tk.StringVar()

# Tạo giao diện để nhập thông tin
tk.Label(root, text="IP Switch:").grid(row=0, column=0, sticky="e")
ip_entry = tk.Entry(root)
ip_entry.grid(row=0, column=1)

tk.Label(root, text="Username:").grid(row=1, column=0, sticky="e")
username_entry = tk.Entry(root)
username_entry.grid(row=1, column=1)

tk.Label(root, text="Password:").grid(row=2, column=0, sticky="e")
password_entry = tk.Entry(root, show="*")
password_entry.grid(row=2, column=1)

tk.Label(root, text="File Excel:").grid(row=3, column=0, sticky="e")
file_entry = tk.Entry(root, textvariable=file_path)
file_entry.grid(row=3, column=1)
browse_button = tk.Button(root, text="Browse", command=browse_file)
browse_button.grid(row=3, column=2)

# Nút thực hiện cấu hình
configure_button = tk.Button(root, text="Cấu hình", command=connect_and_configure)
configure_button.grid(row=4, column=1, sticky="ew")

# Chạy vòng lặp chính của cửa sổ
root.mainloop()
